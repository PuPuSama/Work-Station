from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook


MAX_WORKBOOK_BYTES = 10 * 1024 * 1024
MAX_PREVIEW_COLUMNS = 50
MAX_IMPORT_ROWS = 200
MAX_WORKBOOK_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_WORKBOOK_ARCHIVE_ENTRIES = 5000

IMPORT_FIELDS = (
    "topic",
    "primary_keyword",
    "competitor_keyword",
    "competitor_blog",
)

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "topic": (
        "话题",
        "主题",
        "文章话题",
        "文章主题",
        "文章标题",
        "topic",
        "article topic",
        "title",
    ),
    "primary_keyword": (
        "关键词",
        "主关键词",
        "目标关键词",
        "核心关键词",
        "seo关键词",
        "keyword",
        "primary keyword",
        "target keyword",
        "main keyword",
        "seo keyword",
    ),
    "competitor_keyword": (
        "竞对关键词",
        "竞品关键词",
        "竞争对手关键词",
        "competitor keyword",
        "competitive keyword",
        "competitor keyphrase",
    ),
    "competitor_blog": (
        "竞对blogurl",
        "竞品blogurl",
        "竞争对手blogurl",
        "竞对文章url",
        "竞争对手文章url",
        "竞对链接",
        "竞品链接",
        "blog url",
        "competitor blog",
        "competitor blog url",
        "competitor url",
        "reference url",
    ),
}


class ServerTaskWorkbookError(ValueError):
    """The uploaded workbook cannot be converted into bounded Task rows."""


@dataclass(frozen=True, slots=True)
class ServerTaskWorkbookPreview:
    filename: str
    sheet_name: str
    headers: tuple[str, ...]
    rows: tuple[tuple[str, ...], ...]
    mapping: dict[str, int | None]
    truncated: bool


def _normalized_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).casefold().strip()
    return re.sub(r"[\s_\-—–:：/\\()（）]+", "", normalized)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _field_score(header: str, field: str) -> int:
    normalized = _normalized_header(header)
    if not normalized:
        return 0
    if field == "primary_keyword" and any(
        marker in normalized for marker in ("竞对", "竞品", "竞争对手", "competitor")
    ):
        return 0
    best = 0
    for alias in HEADER_ALIASES[field]:
        normalized_alias = _normalized_header(alias)
        if normalized == normalized_alias:
            best = max(best, 100 + len(normalized_alias))
        elif normalized_alias and normalized_alias in normalized:
            best = max(best, 50 + len(normalized_alias))
    return best


def detect_mapping(headers: tuple[str, ...]) -> dict[str, int | None]:
    mapping: dict[str, int | None] = {field: None for field in IMPORT_FIELDS}
    used: set[int] = set()
    for field in (
        "competitor_blog",
        "competitor_keyword",
        "primary_keyword",
        "topic",
    ):
        candidates = sorted(
            (
                (_field_score(header, field), index)
                for index, header in enumerate(headers)
                if index not in used
            ),
            reverse=True,
        )
        if candidates and candidates[0][0] > 0:
            mapping[field] = candidates[0][1]
            used.add(candidates[0][1])
    return mapping


def _header_score(row: tuple[str, ...]) -> tuple[int, int]:
    mapping = detect_mapping(row)
    detected = sum(index is not None for index in mapping.values())
    return (1 if mapping["topic"] is not None else 0, detected)


def preview_task_workbook(
    *,
    filename: str,
    content: bytes,
) -> ServerTaskWorkbookPreview:
    safe_name = Path(str(filename or "").replace("\\", "/")).name.strip()
    suffix = Path(safe_name).suffix.casefold()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ServerTaskWorkbookError("只支持 .xlsx 或 .xlsm Excel 文件。")
    if not content:
        raise ServerTaskWorkbookError("Excel 文件为空。")
    if len(content) > MAX_WORKBOOK_BYTES:
        raise ServerTaskWorkbookError("Excel 文件超过 10 MB。")
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if (
                len(entries) > MAX_WORKBOOK_ARCHIVE_ENTRIES
                or sum(entry.file_size for entry in entries)
                > MAX_WORKBOOK_UNCOMPRESSED_BYTES
            ):
                raise ServerTaskWorkbookError("Excel 解压后的内容过大。")
    except BadZipFile as exc:
        raise ServerTaskWorkbookError("Excel 文件无法读取或已经损坏。") from exc
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=False,
            data_only=True,
        )
    except Exception as exc:
        raise ServerTaskWorkbookError("Excel 文件无法读取或已经损坏。") from exc
    try:
        best: tuple[tuple[int, int], int, object, list[tuple[str, ...]]] | None = None
        for sheet_index, sheet in enumerate(workbook.worksheets):
            raw_rows: list[tuple[str, ...]] = []
            merged_values: dict[tuple[int, int], str] = {}
            for merged in sheet.merged_cells.ranges:
                if (
                    merged.min_row > 221
                    or merged.min_col > MAX_PREVIEW_COLUMNS
                ):
                    continue
                value = _cell_text(
                    sheet.cell(merged.min_row, merged.min_col).value
                )
                if not value:
                    continue
                for row_number in range(
                    merged.min_row,
                    min(merged.max_row, 221) + 1,
                ):
                    for column_number in range(
                        merged.min_col,
                        min(merged.max_col, MAX_PREVIEW_COLUMNS) + 1,
                    ):
                        merged_values[(row_number, column_number)] = value
            for row_number, raw in enumerate(
                sheet.iter_rows(
                min_row=1,
                max_row=221,
                max_col=MAX_PREVIEW_COLUMNS,
                values_only=True,
                ),
                start=1,
            ):
                row = tuple(
                    _cell_text(value)
                    or merged_values.get((row_number, column_number), "")
                    for column_number, value in enumerate(raw, start=1)
                )
                if any(row):
                    raw_rows.append(row)
            if not raw_rows:
                continue
            for row_index, row in enumerate(raw_rows[:20]):
                score = _header_score(row)
                candidate = (score, -sheet_index, sheet, raw_rows[row_index:])
                if best is None or candidate[:2] > best[:2]:
                    best = candidate
        if best is None:
            raise ServerTaskWorkbookError("Excel 文件中没有可读取的数据行。")
        _score, _sheet_order, selected_sheet, selected_rows = best
        headers = selected_rows[0]
        mapping = detect_mapping(headers)
        data_rows = tuple(
            row
            for row in selected_rows[1 : MAX_IMPORT_ROWS + 1]
            if any(cell.strip() for cell in row)
        )
        if not data_rows:
            raise ServerTaskWorkbookError("识别到表头，但没有可导入的数据行。")
        return ServerTaskWorkbookPreview(
            filename=safe_name,
            sheet_name=str(getattr(selected_sheet, "title", "Sheet")),
            headers=headers,
            rows=data_rows,
            mapping=mapping,
            truncated=len(selected_rows) > MAX_IMPORT_ROWS + 1,
        )
    finally:
        workbook.close()


__all__ = [
    "IMPORT_FIELDS",
    "MAX_WORKBOOK_BYTES",
    "ServerTaskWorkbookError",
    "ServerTaskWorkbookPreview",
    "detect_mapping",
    "preview_task_workbook",
]
