from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config import AppConfig
from models import STATUS_NEW, TaskRecord
from services.task_identity import article_source_key
from storage import now_iso


TOPIC_HINTS = ("话题", "topic", "title")
KEYWORD_HINTS = ("关键词", "竞对", "keyword", "competitor")
BLOG_HINTS = ("url", "blog", "链接")


def normalize_header(value: Any) -> str:
    return str(value or "").replace("\n", "").strip()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def has_any(value: str, hints: tuple[str, ...]) -> bool:
    lower = value.lower()
    return any(hint.lower() in lower for hint in hints)


def find_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    for index, row in enumerate(rows[:20]):
        headers = [normalize_header(value) for value in row]
        joined = " ".join(headers)
        if has_any(joined, TOPIC_HINTS) or has_any(joined, KEYWORD_HINTS) or has_any(joined, BLOG_HINTS):
            return index
    return 0 if rows else None


def find_columns(headers: list[str]) -> tuple[int | None, int | None, int | None]:
    topic_index = None
    keyword_index = None
    blog_index = None

    for index, header in enumerate(headers):
        if has_any(header, TOPIC_HINTS):
            topic_index = index
            break

    for index, header in enumerate(headers):
        if has_any(header, KEYWORD_HINTS):
            keyword_index = index
            break

    for index, header in enumerate(headers):
        if has_any(header, BLOG_HINTS):
            blog_index = index
            break

    if topic_index is None:
        excluded = {index for index in (keyword_index, blog_index) if index is not None}
        for index, header in enumerate(headers):
            if index not in excluded and not header:
                topic_index = index
                break
        if topic_index is None:
            for index in range(len(headers)):
                if index not in excluded:
                    topic_index = index
                    break

    if keyword_index == topic_index:
        keyword_index = None
    if blog_index == topic_index:
        blog_index = None

    return topic_index, keyword_index, blog_index


def make_task_id(customer: str, topic_index: int, topic: str) -> str:
    """Return a stable task id which survives date/week folder changes."""

    return article_source_key(customer, topic, topic_index)[:12]


def scan_topic_library(config: AppConfig) -> list[TaskRecord]:
    config.current_week_path.mkdir(parents=True, exist_ok=True)
    tasks: list[TaskRecord] = []

    for workbook_path in sorted(config.topic_library.glob("*.xlsx")):
        if workbook_path.name.startswith("~$"):
            continue
        customer = workbook_path.stem
        customer_dir = config.current_week_path / customer
        customer_dir.mkdir(parents=True, exist_ok=True)

        article_number = 0
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)

        for sheet in workbook.worksheets:
            sheet.reset_dimensions()
            rows = list(sheet.iter_rows(values_only=True))
            header_row_index = find_header_row(rows)
            if header_row_index is None:
                continue

            headers = [normalize_header(value) for value in rows[header_row_index]]
            topic_index, keyword_index, blog_index = find_columns(headers)
            if topic_index is None:
                continue

            for excel_row_number, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
                topic = cell_text(row[topic_index] if topic_index < len(row) else "")
                if not topic:
                    continue

                article_number += 1
                competitor_keyword = ""
                competitor_blog = ""
                if keyword_index is not None and keyword_index < len(row):
                    competitor_keyword = cell_text(row[keyword_index])
                if blog_index is not None and blog_index < len(row):
                    competitor_blog = cell_text(row[blog_index])

                task_dir = customer_dir / f"topic_{article_number:03d}"
                task_dir.mkdir(parents=True, exist_ok=True)
                task_id = make_task_id(customer, article_number, topic)
                created_at = now_iso()
                task = TaskRecord(
                    id=task_id,
                    week_folder=config.current_week_folder,
                    customer=customer,
                    source_key=article_source_key(customer, topic, article_number),
                    topic_index=article_number,
                    topic=topic,
                    competitor_keyword=competitor_keyword,
                    competitor_blog=competitor_blog,
                    status=STATUS_NEW,
                    task_dir=str(task_dir),
                    created_at=created_at,
                    updated_at=created_at,
                )
                source = {
                    "workbook": str(workbook_path),
                    "sheet": sheet.title,
                    "excel_row_number": excel_row_number,
                    "headers": headers,
                    "topic": topic,
                    "competitor_keyword": competitor_keyword,
                    "competitor_blog": competitor_blog,
                }
                (task_dir / "source_row.json").write_text(
                    json.dumps(source, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                tasks.append(task)

        workbook.close()

    return tasks
