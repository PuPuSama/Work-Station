from __future__ import annotations

import sys
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook


BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from services.server_task_workbook import (  # noqa: E402
    ServerTaskWorkbookError,
    preview_task_workbook,
)


def workbook_bytes(rows: list[list[object]], *, sheet_name: str = "Topics") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class ServerTaskWorkbookTests(unittest.TestCase):
    def test_detects_four_business_columns_without_conflating_keywords(self) -> None:
        preview = preview_task_workbook(
            filename="topics.xlsx",
            content=workbook_bytes(
                [
                    ["说明", "季度选题"],
                    [
                        "文章话题",
                        "目标关键词",
                        "竞对关键词",
                        "竞争对手 Blog URL",
                    ],
                    [
                        "How to source DIN 985 nuts",
                        "DIN 985 nuts",
                        "nylon lock nuts",
                        "https://competitor.example/blog/din-985",
                    ],
                ]
            ),
        )
        self.assertEqual(preview.sheet_name, "Topics")
        self.assertEqual(
            preview.mapping,
            {
                "topic": 0,
                "primary_keyword": 1,
                "competitor_keyword": 2,
                "competitor_blog": 3,
            },
        )
        self.assertEqual(len(preview.rows), 1)
        self.assertEqual(preview.rows[0][0], "How to source DIN 985 nuts")

    def test_selects_the_sheet_with_the_best_header_match(self) -> None:
        workbook = Workbook()
        workbook.active.append(["Notes", "Not an import sheet"])
        sheet = workbook.create_sheet("Import")
        sheet.append(["Topic", "Primary Keyword", "Competitor URL"])
        sheet.append(["Bolt guide", "bolts", "https://example.com/guide"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        preview = preview_task_workbook(
            filename="multi-sheet.xlsx",
            content=output.getvalue(),
        )
        self.assertEqual(preview.sheet_name, "Import")
        self.assertEqual(preview.mapping["topic"], 0)
        self.assertEqual(preview.mapping["primary_keyword"], 1)
        self.assertEqual(preview.mapping["competitor_blog"], 2)

    def test_expands_a_merged_keyword_only_across_its_excel_range(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["话题", "关键词"])
        sheet.append(["Topic A", "hybrid inverter"])
        sheet.append(["Topic B", None])
        sheet.append(["Topic C", None])
        sheet.append(["Topic D", None])
        sheet.append(["Topic E", None])
        sheet.merge_cells("B2:B4")
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        preview = preview_task_workbook(
            filename="merged-keywords.xlsx",
            content=output.getvalue(),
        )
        keyword_column = preview.mapping["primary_keyword"]
        assert keyword_column is not None
        self.assertEqual(
            [row[keyword_column] for row in preview.rows],
            ["hybrid inverter", "hybrid inverter", "hybrid inverter", "", ""],
        )

    def test_rejects_legacy_xls_and_oversized_payloads(self) -> None:
        with self.assertRaisesRegex(ServerTaskWorkbookError, "xlsx"):
            preview_task_workbook(filename="topics.xls", content=b"xls")
        with self.assertRaisesRegex(ServerTaskWorkbookError, "10 MB"):
            preview_task_workbook(
                filename="topics.xlsx",
                content=b"x" * (10 * 1024 * 1024 + 1),
            )


if __name__ == "__main__":
    unittest.main()
